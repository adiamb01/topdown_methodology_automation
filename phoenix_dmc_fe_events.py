#!/usr/bin/env python3
"""
Phoenix DMC FE metric extractor.

This script can either:
  1) run topdown-tool for DMC frontend metrics/events, or
  2) parse an existing topdown output directory.

It extracts the requested DMC FE metrics, applies formulas from the raw
frontend event CSV, and writes an Excel workbook with a Summary tab.

Typical run:
  python3 phoenix_dmc_fe_summary.py \
    --tool /root/aambre/telemetry-solution/tools/topdown_tool/.venv/bin/topdown-tool \
    --out dmc_fe_out \
    --workload "taskset -c 4 bw_mem -P 1 -N 1000 -W 2 1024m rd"

Parse existing output:
  python3 phoenix_dmc_fe_summary.py \
    --input-dir /root/aambre/telemetry-solution/tools/topdown_tool/dmc_missing_metrics/2026_06_28_15_52_50 \
    --out dmc_fe_summary \
    --chip 0

Chip selection:
  --chip 0     lower-half DMC IDs
  --chip 1     upper-half DMC IDs
  --chip both  all DMC IDs
"""

from __future__ import annotations

import argparse
import csv
import math
import os
import re
import shlex
import subprocess
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    import pandas as pd
except Exception as exc:
    raise SystemExit(f"[ERROR] pandas is required: {exc}")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception as exc:
    raise SystemExit(f"[ERROR] openpyxl is required: python3 -m pip install openpyxl. Details: {exc}")


DEFAULT_GROUPS = [
    "IF_Command_Mix",
    "IF_BW_Analysis",
    "IF_Load_Analysis",
    "IF_Retry_Analysis",
    "IF_Zero_Credits",
    "Queue_Effectiveness",
]

REQUESTED_METRICS = [
    "dmc_fe_req_rate",
    "dmc_fe_request_stall",
    "dmc_fe_response_stall",
    "dmc_fe_write_request_util",
    "dmc_fe_write_retry_util",
    "dmc_fe_writenosnp_pcmosep_rate",
    "dmc_fe_writenosnpfull_rate",
    "dmc_fe_writenosnpptl_rate",
    "dmc_fe_writenosnpzero_rate",
    "DMC_FE_CMD_TYPE_CYCLES",
    "DMC_FE_CYCLES",
]

# Tool-native metric names that may already be present in dmc_frontend_metrics.csv.
TOOL_METRIC_ALIASES = {
    "dmc_fe_req_rate": ["chi_req_if_rcv_percentage"],
    "dmc_fe_request_stall": ["chi_request_zerocredits"],
    "dmc_fe_response_stall": ["chi_response_zerocredits"],
    "dmc_fe_write_request_util": ["chi_wr_percentage"],
    "dmc_fe_write_retry_util": ["chi_retry_percentage"],
    "dmc_fe_writenosnp_pcmosep_rate": ["chi_writenosnp_pcmosep_percentage"],
    "dmc_fe_writenosnpfull_rate": ["chi_writenosnpfull_percentage"],
    "dmc_fe_writenosnpptl_rate": ["chi_writenosnpptl_percentage"],
    "dmc_fe_writenosnpzero_rate": ["chi_writenosnpzero_percentage"],
}

# Formula event mappings.
CYCLES = "CHI_CYCLES"
EVENT_FORMULAS = {
    # Phoenix FE telemetry-style formulas using raw frontend event counts.
    # *_rate command-mix metrics are normalized to total command-type cycles.
    # *_util / *_stall metrics are normalized to FE cycles.
    "dmc_fe_request_stall": {
        "num": ["CHI_REQIN_TIMEOUT_CREDIT"],
        "den": CYCLES,
        "unit": "%",
        "formula": "CHI_REQIN_TIMEOUT_CREDIT / CHI_CYCLES * 100",
    },
    "dmc_fe_response_stall": {
        "num": ["CHI_RSPOUT_TIMEOUT_CREDIT"],
        "den": CYCLES,
        "unit": "%",
        "formula": "CHI_RSPOUT_TIMEOUT_CREDIT / CHI_CYCLES * 100",
    },
    # Use tool-derived chi_retry_percentage for this metric name.
    # Do not derive from CHI_REQ_XMIT_WR_RETRIES because Phoenix topdown exports
    # the expected ~40% retry value as chi_retry_percentage.

}


CMD_TYPE_EVENTS = [
    "CHI_REQIF_OP_CLEANSHAREDPERSIST",
    "CHI_REQIF_OP_PREFETCHTGT",
    "CHI_REQIF_OP_READNOSNP",
    "CHI_REQIF_OP_READNOSNPSEP",
    "CHI_REQIF_OP_WRITENOSNPFULL",
    "CHI_REQIF_OP_WRITENOSNPFULL_PTL_PCMOSEP",
    "CHI_REQIF_OP_WRITENOSNPPTL",
    "CHI_REQIF_OP_WRITEZERO",
]

WRITE_REQUEST_EVENTS = [
    "CHI_REQIF_OP_WRITENOSNPFULL",
    "CHI_REQIF_OP_WRITENOSNPFULL_PTL_PCMOSEP",
    "CHI_REQIF_OP_WRITENOSNPPTL",
    "CHI_REQIF_OP_WRITEZERO",
]


def numeric_device_cols(df: pd.DataFrame, prefix: str = "frontend") -> List[str]:
    cols = []
    for c in df.columns:
        if c.startswith(prefix) and c.endswith("-value"):
            cols.append(c)
    return cols


def safe_float(x) -> Optional[float]:
    try:
        if x is None:
            return None
        s = str(x).strip()
        if not s or s.lower() in {"nan", "nan.0", "nan.000000", "<not counted>", "<not supported>"}:
            return None
        v = float(s)
        if math.isnan(v):
            return None
        return v
    except Exception:
        return None


def device_index(dev: str) -> Optional[int]:
    m = re.search(r"(?:frontend|backend)(\d+)", str(dev))
    return int(m.group(1)) if m else None


def selected_devices(all_devices: List[str], chip: str) -> List[str]:
    """Return frontend/backend device labels selected by chip.

    Phoenix DMC frontend IDs are sparse:
      chip 0: frontend0..frontend11
      chip 1: frontend24..frontend35

    topdown CSVs may include NaN columns for the other chip, so do not split by
    midpoint of min/max device IDs.
    """
    chip = str(chip).lower()
    devs = sorted(set(all_devices), key=lambda d: (device_index(d) is None, device_index(d) or -1, d))
    if chip in {"both", "all"}:
        return devs

    selected = []
    for d in devs:
        i = device_index(d)
        if i is None:
            continue
        if chip == "0" and 0 <= i <= 11:
            selected.append(d)
        elif chip == "1" and 24 <= i <= 35:
            selected.append(d)

    if chip not in {"0", "1"}:
        raise SystemExit("[ERROR] --chip must be 0, 1, or both")

    return selected


def filter_device_map(vals: Dict[str, float], chip: str) -> Dict[str, float]:
    keep = set(selected_devices(list(vals.keys()), chip))
    return {d: v for d, v in vals.items() if d in keep}


def find_latest_capture(out_root: Path) -> Path:
    if (out_root / "dmc").exists():
        return out_root
    dirs = [p for p in out_root.iterdir() if p.is_dir()]
    if not dirs:
        return out_root
    # Prefer dirs that contain dmc output.
    dmc_dirs = [p for p in dirs if (p / "dmc").exists()]
    if dmc_dirs:
        return sorted(dmc_dirs, key=lambda p: p.stat().st_mtime)[-1]
    return sorted(dirs, key=lambda p: p.stat().st_mtime)[-1]


def run_topdown(args: argparse.Namespace) -> Path:
    out_root = Path(args.out).resolve()
    out_root.mkdir(parents=True, exist_ok=True)

    cmd = [
        args.tool,
        "--probe", "DMC",
        "--hwplatform", args.hwplatform,
        "--dmc-generate-csv", "metrics,events",
        "--csv-output-path", str(out_root),
    ]

    if not args.no_groups:
        groups = args.dmc_metric_group or ",".join(DEFAULT_GROUPS)
        cmd += ["--dmc-metric-group", groups]

    if args.extra_topdown_args:
        cmd += shlex.split(args.extra_topdown_args)

    if not args.workload:
        raise SystemExit("[ERROR] --workload is required unless --input-dir is used")

    cmd.append("--")
    cmd += shlex.split(args.workload)

    log = out_root / "topdown_run.log"
    print("[INFO] running:", shlex.join(cmd))
    with log.open("w", encoding="utf-8") as f:
        f.write("COMMAND: " + shlex.join(cmd) + "\n\n")
        rc = subprocess.run(cmd, stdout=f, stderr=subprocess.STDOUT, text=True, check=False).returncode
        f.write(f"\nEXIT_STATUS: {rc}\n")
    if rc != 0:
        print(f"[WARN] topdown-tool exited with rc={rc}. See {log}", file=sys.stderr)

    capture = find_latest_capture(out_root)
    return capture


def load_frontend_csvs(capture_dir: Path) -> Tuple[pd.DataFrame, pd.DataFrame]:
    candidates = [
        capture_dir / "dmc",
        capture_dir,
    ]
    events_path = metrics_path = None
    for d in candidates:
        e = d / "dmc_frontend_events.csv"
        m = d / "dmc_frontend_metrics.csv"
        if e.exists():
            events_path = e
        if m.exists():
            metrics_path = m
    if events_path is None:
        found = list(capture_dir.rglob("dmc_frontend_events.csv"))
        if found:
            events_path = found[0]
    if metrics_path is None:
        found = list(capture_dir.rglob("dmc_frontend_metrics.csv"))
        if found:
            metrics_path = found[0]

    if events_path is None:
        raise SystemExit(f"[ERROR] could not find dmc_frontend_events.csv under {capture_dir}")
    if metrics_path is None:
        raise SystemExit(f"[ERROR] could not find dmc_frontend_metrics.csv under {capture_dir}")

    print(f"[INFO] events : {events_path}")
    print(f"[INFO] metrics: {metrics_path}")
    return pd.read_csv(events_path), pd.read_csv(metrics_path)


def event_map(events_df: pd.DataFrame) -> Dict[str, Dict[str, float]]:
    dev_cols = numeric_device_cols(events_df)
    out: Dict[str, Dict[str, float]] = {}
    for _, row in events_df.iterrows():
        ev = str(row.get("event", "")).strip()
        if not ev:
            continue
        vals: Dict[str, float] = {}
        for c in dev_cols:
            v = safe_float(row.get(c))
            if v is not None:
                vals[c.replace("-value", "")] = v
        out[ev] = vals
    return out


def metric_tool_values(metrics_df: pd.DataFrame, metric_name: str) -> Dict[str, float]:
    """Return values for a tool-native metric alias.

    Prefer per-device frontend*-value columns when present. If the metrics CSV
    only has aggregate/average columns, fall back to the row average and expose
    it as a synthetic "tool_average" device. This keeps the output metric name
    unchanged while allowing dmc_fe_write_retry_util to source
    chi_retry_percentage.
    """
    aliases = TOOL_METRIC_ALIASES.get(metric_name, [])
    if not aliases or "metric" not in metrics_df.columns:
        return {}

    sub = metrics_df[metrics_df["metric"].astype(str).isin(aliases)]
    if sub.empty:
        return {}

    row = sub.iloc[0]
    vals: Dict[str, float] = {}

    dev_cols = numeric_device_cols(metrics_df)
    for c in dev_cols:
        v = safe_float(row.get(c))
        if v is not None:
            vals[c.replace("-value", "")] = v

    if vals:
        return vals

    # Fallback for tool-derived rows with only summary columns.
    for c in ("average", "Average", "avg", "AVG", "value", "Value"):
        if c in metrics_df.columns:
            v = safe_float(row.get(c))
            if v is not None:
                return {"tool_average": v}

    return {}


def divide_device_maps(num_maps: List[Dict[str, float]], den_map: Dict[str, float], scale: float) -> Dict[str, float]:
    devices = set(den_map)
    for m in num_maps:
        devices |= set(m)
    out = {}
    for dev in sorted(devices):
        den = den_map.get(dev)
        if den is None or den == 0:
            continue
        num = sum(m.get(dev, 0.0) for m in num_maps)
        out[dev] = num / den * scale
    return out


def summarize_values(vals: Dict[str, float]) -> Tuple[Optional[float], Optional[float], Optional[float], Optional[float], int]:
    if not vals:
        return None, None, None, None, 0
    xs = list(vals.values())
    return sum(xs), sum(xs) / len(xs), min(xs), max(xs), len(xs)


def sum_event_maps(maps: List[Dict[str, float]]) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for m in maps:
        for dev, value in m.items():
            out[dev] = out.get(dev, 0.0) + value
    return out


def ratio_map(num_map: Dict[str, float], den_map: Dict[str, float], scale: float = 100.0) -> Dict[str, float]:
    out: Dict[str, float] = {}
    for dev, num in num_map.items():
        den = den_map.get(dev)
        if den is None or den == 0:
            continue
        out[dev] = num / den * scale
    return out


def compute_summary(events_df: pd.DataFrame, metrics_df: pd.DataFrame, chip: str = "both") -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    evs = event_map(events_df)

    cycles_map = evs.get(CYCLES, {})
    cmd_type_map = sum_event_maps([evs[e] for e in CMD_TYPE_EVENTS if e in evs])
    write_cmd_events = [
        "CHI_REQIF_OP_WRITENOSNPFULL",
        "CHI_REQIF_OP_WRITENOSNPPTL",
        "CHI_REQIF_OP_WRITEZERO",
        "CHI_REQIF_OP_WRITENOSNPFULL_PTL_PCMOSEP",
    ]
    write_cmd_map = sum_event_maps([evs[e] for e in write_cmd_events if e in evs])

    computed: Dict[str, Tuple[Dict[str, float], str, str, str]] = {}
    computed["DMC_FE_CYCLES"] = (cycles_map, "cycles", "raw event", "CHI_CYCLES")
    computed["DMC_FE_CMD_TYPE_CYCLES"] = (cmd_type_map, "cycles", "raw events", "sum(CHI_REQIF_OP_*)")
    computed["dmc_fe_req_rate"] = (
        ratio_map(cmd_type_map, cycles_map),
        "%",
        "computed from events",
        "sum(CHI_REQIF_OP_*) / CHI_CYCLES * 100",
    )
    computed["dmc_fe_write_request_util"] = (
        ratio_map(write_cmd_map, cycles_map),
        "%",
        "computed from events",
        "sum(write CHI_REQIF_OP_*) / CHI_CYCLES * 100",
    )
    # Command mix metrics: denominator is total command-type cycles, not FE cycles.
    command_mix_map = {
        "dmc_fe_writenosnp_pcmosep_rate": "CHI_REQIF_OP_WRITENOSNPFULL_PTL_PCMOSEP",
        "dmc_fe_writenosnpfull_rate": "CHI_REQIF_OP_WRITENOSNPFULL",
        "dmc_fe_writenosnpptl_rate": "CHI_REQIF_OP_WRITENOSNPPTL",
        "dmc_fe_writenosnpzero_rate": "CHI_REQIF_OP_WRITEZERO",
    }
    for metric_name, event_name in command_mix_map.items():
        computed[metric_name] = (
            ratio_map(evs.get(event_name, {}), cmd_type_map),
            "%",
            "computed from events",
            f"{event_name} / sum(CHI_REQIF_OP_*) * 100",
        )

    for metric_name, f in EVENT_FORMULAS.items():
        missing = [e for e in f["num"] + [f["den"]] if e not in evs]
        if missing:
            vals = metric_tool_values(metrics_df, metric_name)
            if vals:
                computed[metric_name] = (vals, "%", "tool metric fallback", ",".join(TOOL_METRIC_ALIASES.get(metric_name, [])))
            else:
                computed[metric_name] = ({}, "", "missing", "missing raw event(s): " + ",".join(missing))
        else:
            computed[metric_name] = (
                divide_device_maps([evs[e] for e in f["num"]], evs[f["den"]], 100.0),
                f["unit"],
                "computed from events",
                f["formula"],
            )

    # Keep output metric name as dmc_fe_write_retry_util, but source the value
    # from topdown's chi_retry_percentage. Do not use aggregate; use per-device
    # frontend*-value columns and then average/min/max over the selected chip.
    retry_vals = metric_tool_values(metrics_df, "dmc_fe_write_retry_util")
    computed["dmc_fe_write_retry_util"] = (
        retry_vals,
        "%" if retry_vals else "",
        "tool metric",
        "chi_retry_percentage",
    )

    summary_rows = []
    per_device_rows = []
    for metric in REQUESTED_METRICS:
        vals, unit, source, formula = computed.get(metric, ({}, "", "missing", ""))
        vals = filter_device_map(vals, chip)
        total, avg, mn, mx, n = summarize_values(vals)
        summary_rows.append({
            "metric": metric,
            "value": avg,
            "min": mn,
            "max": mx,
            "unit": unit,
        })
        for dev, v in vals.items():
            per_device_rows.append({
                "metric": metric,
                "chip": chip,
                "device": dev,
                "value": v,
                "unit": unit,
                "source": source,
                "formula": formula,
            })

    # Tool metric alias table for comparison/debug only. The Summary tab uses
    # the explicit event formulas above.
    tool_rows = []
    for requested, aliases in TOOL_METRIC_ALIASES.items():
        for alias in aliases:
            if "metric" not in metrics_df.columns:
                continue
            sub = metrics_df[metrics_df["metric"].astype(str).eq(alias)]
            if sub.empty:
                continue
            vals = metric_tool_values(metrics_df, requested)
            total, avg, mn, mx, n = summarize_values(vals)
            tool_rows.append({
                "requested_metric": requested,
                "tool_metric": alias,
                "device_count": n,
                "sum": total,
                "average": avg,
                "min": mn,
                "max": mx,
            })

    return pd.DataFrame(summary_rows, columns=["metric", "value", "min", "max", "unit"]), pd.DataFrame(per_device_rows), pd.DataFrame(tool_rows)

def write_excel(path: Path, summary: pd.DataFrame, per_device: pd.DataFrame, tool_metrics: pd.DataFrame, events_df: pd.DataFrame, metrics_df: pd.DataFrame) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    def write_df(sheet, df: pd.DataFrame):
        sheet.append(list(df.columns))
        for _, row in df.iterrows():
            sheet.append([row.get(c) for c in df.columns])
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sheet.freeze_panes = "A2"
        for i, col in enumerate(df.columns, 1):
            max_len = max([len(str(col))] + [len(str(v)) for v in df[col].head(200).tolist()])
            sheet.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 70)

    write_df(ws, summary)

    ws2 = wb.create_sheet("Per_Device")
    write_df(ws2, per_device)

    ws3 = wb.create_sheet("Tool_Metric_Aliases")
    write_df(ws3, tool_metrics if not tool_metrics.empty else pd.DataFrame(columns=["requested_metric", "tool_metric", "device_count", "sum", "average", "min", "max"]))

    # Keep raw CSVs in workbook too.
    ws4 = wb.create_sheet("Raw_Frontend_Events")
    write_df(ws4, events_df)

    ws5 = wb.create_sheet("Raw_Frontend_Metrics")
    write_df(ws5, metrics_df)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Run/parse Phoenix DMC FE metrics and create Excel summary")
    ap.add_argument("--tool", default="topdown-tool", help="Path to topdown-tool")
    ap.add_argument("--hwplatform", default="phoenix")
    ap.add_argument("--out", default="dmc_fe_summary_out", help="Output directory")
    ap.add_argument("--workload", default=None, help='Workload command, e.g. "taskset -c 4 bw_mem -P 1 -N 1000 -W 2 1024m rd"')
    ap.add_argument("--input-dir", default=None, help="Existing topdown capture directory to parse instead of running")
    ap.add_argument("--dmc-metric-group", default=None, help="Comma-separated DMC metric groups. Default uses IF frontend groups.")
    ap.add_argument("--no-groups", action="store_true", help="Do not pass --dmc-metric-group; let topdown-tool collect defaults")
    ap.add_argument("--extra-topdown-args", default="", help="Extra args passed before --")
    ap.add_argument("--chip", default="both", choices=["0", "1", "both", "all"], help="DMC chip selection: 0=lower-half DMC IDs, 1=upper-half DMC IDs, both=all DMC IDs")
    ap.add_argument("--xlsx", default=None, help="Output Excel path. Default: <out>/dmc_fe_metric_summary.xlsx")
    args = ap.parse_args()

    out_dir = Path(args.out).resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.input_dir:
        capture_dir = Path(args.input_dir).resolve()
    else:
        capture_dir = run_topdown(args)

    events_df, metrics_df = load_frontend_csvs(capture_dir)
    summary, per_device, tool_metrics = compute_summary(events_df, metrics_df, chip=args.chip)

    xlsx = Path(args.xlsx).resolve() if args.xlsx else out_dir / "dmc_fe_metric_summary.xlsx"
    write_excel(xlsx, summary, per_device, tool_metrics, events_df, metrics_df)

    summary_csv = out_dir / "dmc_fe_metric_summary.csv"
    per_device_csv = out_dir / "dmc_fe_metric_per_device.csv"
    summary.to_csv(summary_csv, index=False)
    per_device.to_csv(per_device_csv, index=False)

    print(f"\n=== Summary chip={args.chip} ===")
    cols = ["metric", "value", "min", "max", "unit"]
    print(summary[cols].to_string(index=False))
    print(f"\nWrote: {xlsx}")
    print(f"Wrote: {summary_csv}")
    print(f"Wrote: {per_device_csv}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
